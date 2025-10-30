import os
import asyncio
from datetime import datetime
import openpyxl
import sqlite3
from contextlib import closing
import aiohttp
from aiogram import Bot, Dispatcher, Router, F
from aiogram.types import (
    Message, CallbackQuery,
    ReplyKeyboardMarkup, KeyboardButton,
    InlineKeyboardMarkup, InlineKeyboardButton,
    InputMediaPhoto, InputMediaVideo, InputMediaDocument, BotCommand
)
from aiogram.filters import Command
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.exceptions import TelegramRetryAfter, TelegramBadRequest
from aiohttp import web
import json
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ========== НАСТРОЙКИ ==========
TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN")
WEBHOOK_URL = os.environ.get("WEBHOOK_URL", "https://telegram-bot-b6pn.onrender.com")
PORT = int(os.environ.get("PORT", 10000))
DB_PATH = "files.db"

bot = Bot(token=TOKEN)
storage = MemoryStorage()
dp = Dispatcher(storage=storage)
router = Router()

# ========== МАРШРУТИЗАЦИЯ ТЕМ (WORK → ARCHIVE) ==========
# Оставлены настройки в коде; вы изменяете их программно
TOPIC_MAP = {
    -1003281117256: {  # Рабочая группа A (dagestan.xlsx)
        3: {"chat_id": -1003250982118, "thread_id": 3},
    },
    -1003237477689: {  # Рабочая группа B (nazran.xlsx)
        15: {"chat_id": -1003252316518, "thread_id": 6},
    },
    -1004000000001: {i: {"chat_id": -1005000000001, "thread_id": i} for i in range(1, 11)},
    -1004000000002: {i: {"chat_id": -1005000000002, "thread_id": i} for i in range(1, 11)},
    -1004000000003: {i: {"chat_id": -1005000000003, "thread_id": i} for i in range(1, 11)},
}

# ========== ПРИВЯЗКА EXCEL К РАБОЧИМ ГРУППАМ ==========
EXCEL_MAP = {
    -1003281117256: "dagestan.xlsx",
    -1003237477689: "nazran.xlsx",
    -1004000000001: "bryunsk.xlsx",
    -1004000000002: "orel.xlsx",
    -1004000000003: "objects.xlsx",
}

# ========== КОНСТАНТЫ ==========
UPLOAD_STEPS = [
    "📸 Общий вид газопровода до и после счётчика",
    "🧾 Старый счётчик — общий и крупный план (заводской номер, год, показания)",
    "🔒 Фото пломбы на фоне маркировки счётчика",
    "➡️ Стрелка направления газа на старом счётчике",
    "🧱 Газопровод после монтажа нового счётчика",
    "🆕 Новый счётчик — общий и крупный план (заводской номер, год, показания)",
    "➡️ Стрелка направления нового счётчика",
    "🔥 Шильдик котла и других приборов (модель и мощность)",
]
VIDEO_STEP = "🎥 Видео герметичности соединений"

# ========== БАЗА ДАННЫХ ==========
def init_db():
    with closing(sqlite3.connect(DB_PATH)) as conn:
        conn.execute("""CREATE TABLE IF NOT EXISTS files(\
            id INTEGER PRIMARY KEY AUTOINCREMENT,\
            object_id TEXT, step TEXT, kind TEXT, file_id TEXT,\
            author TEXT, created_at TEXT\
        )""")
        conn.execute("""CREATE TABLE IF NOT EXISTS completed(\
            object_id TEXT PRIMARY KEY,\
            author TEXT,\
            completed_at TEXT\
        )""")
        conn.execute("""CREATE TABLE IF NOT EXISTS settings(\
            key TEXT PRIMARY KEY,\
            value TEXT\
        )""")
        conn.commit()

def db_set(key, value):
    with closing(sqlite3.connect(DB_PATH)) as conn:
        conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, value))
        conn.commit()

def db_get(key):
    with closing(sqlite3.connect(DB_PATH)) as conn:
        cur = conn.execute("SELECT value FROM settings WHERE key=?", (key,))
        r = cur.fetchone()
        return r[0] if r else None

def save_files(object_id, step, files, author):
    if not files:
        return
    with closing(sqlite3.connect(DB_PATH)) as conn:
        conn.executemany(
            "INSERT INTO files(object_id, step, kind, file_id, author, created_at) VALUES (?,?,?,?,?,?)",
            [(object_id, step, f["type"], f["file_id"], author, datetime.now().isoformat()) for f in files]
        )
        conn.commit()

def get_files(object_id):
    with closing(sqlite3.connect(DB_PATH)) as conn:
        cur = conn.execute("SELECT step, kind, file_id FROM files WHERE object_id=? ORDER BY id", (object_id,))
        data = {}
        for step, kind, file_id in cur.fetchall():
            data.setdefault(step, []).append({"type": kind, "file_id": file_id})
        return data

def delete_files_by_object(object_id):
    with closing(sqlite3.connect(DB_PATH)) as conn:
        conn.execute("DELETE FROM files WHERE object_id=?", (object_id,))
        conn.commit()

def mark_completed(object_id: str, author: str):
    with closing(sqlite3.connect(DB_PATH)) as conn:
        conn.execute(
            "INSERT OR REPLACE INTO completed(object_id, author, completed_at) VALUES (?,?,?)",
            (object_id, author, datetime.now().isoformat())
        )
        conn.commit()

def list_completed():
    with closing(sqlite3.connect(DB_PATH)) as conn:
        cur = conn.execute("SELECT object_id, author, completed_at FROM completed ORDER BY object_id")
        return cur.fetchall()

# ========== HELPERS ==========
def stringify_keys(d):
    if isinstance(d, dict):
        return {str(k): stringify_keys(v) for k, v in d.items()}
    elif isinstance(d, list):
        return [stringify_keys(x) for x in d]
    else:
        return d

def load_settings():
    global TOPIC_MAP, EXCEL_MAP
    try:
        val1 = db_get("TOPIC_MAP")
        val2 = db_get("EXCEL_MAP")
        if val1:
            TOPIC_MAP = json.loads(val1)
        if val2:
            EXCEL_MAP = json.loads(val2)
        logger.info("Настройки загружены из БД.")
    except Exception as e:
        logger.exception("Ошибка загрузки настроек: %s", e)

def save_settings():
    db_set("TOPIC_MAP", json.dumps(stringify_keys(TOPIC_MAP), ensure_ascii=False))
    db_set("EXCEL_MAP", json.dumps(stringify_keys(EXCEL_MAP), ensure_ascii=False))

# ========== STATES ==========
class Upload(StatesGroup):
    waiting_object = State()
    confirming = State()
    uploading = State()
    uploading_video = State()

class AddPhoto(StatesGroup):
    waiting_object = State()
    confirming = State()
    uploading = State()

class Info(StatesGroup):
    waiting_object = State()

# ========== KEYBOARDS ==========
def main_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="/addphoto"), KeyboardButton(text="/info"), KeyboardButton(text="/photo")]
        ],
        resize_keyboard=True
    )

def cancel_only_kb(user_id: int | None):
    cancel_cb = f"cancel_{user_id}" if user_id else "cancel"
    return InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="❌ Отмена", callback_data=cancel_cb)]
    ])

def action_kb(user_id: int | None):
    uid = user_id or ""
    return InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="💾 Сохранить", callback_data=f"save_{uid}"),
        InlineKeyboardButton(text="❌ Отмена", callback_data=f"cancel_{uid}")
    ]])

def confirm_kb(prefix: str, user_id: int):
    return InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Да", callback_data=f"{prefix}_confirm_yes_{user_id}"),
        InlineKeyboardButton(text="❌ Отмена", callback_data=f"{prefix}_confirm_no_{user_id}")
    ]])

# ========== SAFE CALLS ==========
async def safe_call(coro, pause=0.25):
    try:
        res = await coro
        await asyncio.sleep(pause)
        return res
    except TelegramRetryAfter as e:
        await asyncio.sleep(e.retry_after + 1)
        return await coro
    except TelegramBadRequest as e:
        logger.debug("Ignored TelegramBadRequest in safe_call: %s", e)
        return None
    except Exception:
        logger.exception("Unexpected error in safe_call")
        return None

async def safe_cq_answer(cq: CallbackQuery, text: str | None = None, show_alert: bool = False):
    try:
        if text:
            await cq.answer(text, show_alert=show_alert)
        else:
            await cq.answer()
    except TelegramBadRequest as e:
        logger.debug("Ignored TelegramBadRequest in cq.answer: %s", e)
    except Exception:
        logger.exception("Unexpected error in safe_cq_answer")

async def safe_edit_message(msg, *args, **kwargs):
    try:
        await msg.edit_text(*args, **kwargs)
    except TelegramBadRequest as e:
        logger.debug("Ignored TelegramBadRequest in edit_text: %s", e)
    except Exception:
        logger.exception("Unexpected error editing message")

# ========== KEEPALIVE ==========
async def keepalive():
    while True:
        try:
            async with aiohttp.ClientSession() as s:
                await s.get(WEBHOOK_URL)
        except Exception:
            logger.debug("keepalive failed", exc_info=True)
        await asyncio.sleep(240)

# ========== HELPERS FOR EXCEL / MAPPING ==========
def get_excel_filename_for_chat(chat_id: int) -> str | None:
    return EXCEL_MAP.get(str(chat_id)) or EXCEL_MAP.get(chat_id)

def mapping_lookup(work_chat_id: int | str, work_thread_id: int | str):
    sub = TOPIC_MAP.get(work_chat_id) or TOPIC_MAP.get(str(work_chat_id))
    if not sub:
        return None
    return sub.get(work_thread_id) or sub.get(str(work_thread_id))

def check_object_excel(chat_id: int, object_id: str):
    filename = get_excel_filename_for_chat(chat_id)
    if not filename:
        return None, "⚠️ К этой группе не привязан Excel-документ. Обратитесь к администратору."
    try:
        wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        sh = wb.active
        for row in sh.iter_rows(min_row=2, values_only=True):
            if str(row[0]).strip() == str(object_id):
                return True, str(row[1])
        return False, None
    except Exception as e:
        return None, f"{filename}: {e}"

def get_object_info(chat_id: int, object_id: str):
    filename = get_excel_filename_for_chat(chat_id)
    if not filename:
        return {"error": "⚠️ К этой группе не привязан Excel-документ. Обратитесь к администратору."}
    try:
        wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        sh = wb.active
        for row in sh.iter_rows(min_row=2, values_only=True):
            if str(row[0]).strip() == str(object_id):
                return {
                    "id": str(row[0]).strip(),
                    "consumer": str(row[1]) if len(row) > 1 else "Н/Д",
                    "object": str(row[2]) if len(row) > 2 else "Н/Д",
                    "address": str(row[3]) if len(row) > 3 else "Н/Д",
                }
        return None
    except Exception as e:
        return {"error": f"{filename}: {e}"}

# проверка строгой привязки: команда может выполняться ТОЛЬКО в теме (message_thread_id),
# и только если для этой темы есть маршрут в TOPIC_MAP
def is_topic_allowed(chat_id: int, thread_id: int | None) -> bool:
    if thread_id is None:
        return False
    return mapping_lookup(chat_id, thread_id) is not None

# ========== USER COMMANDS ==========
@router.message(Command("start"))
async def cmd_start(m: Message):
    await m.answer(
        "👋 Привет! Это бот для загрузки фото по объектам счётчиков газа.\n\n"
        "📎 /addphoto — добавить фото\n"
        "ℹ️ /info — информация по объектам\n"
        "📸 /photo — новая загрузка\n"
        "⚙️ Работает в темах форумах (супергруппы).",
        reply_markup=main_kb()
    )

@router.message(Command("addphoto"))
async def cmd_addphoto(m: Message, state: FSMContext):
    # строгая привязка к теме
    thread_id = getattr(m, "message_thread_id", None)
    if not is_topic_allowed(m.chat.id, thread_id):
        await m.answer("⚠️ Эта команда доступна только в конкретных темах (топиках). Обратитесь к администратору.")
        return
    await state.set_state(AddPhoto.waiting_object)
    await state.update_data(
        owner_id=m.from_user.id,
        work_chat_id=m.chat.id,
        work_thread_id=thread_id,
        step_msg=None,
        files=[]
    )
    await m.answer("🔢 Введите номер объекта (для добавления файлов):")

@router.message(Command("info"))
async def cmd_info(m: Message, state: FSMContext):
    # строгая привязка к теме
    thread_id = getattr(m, "message_thread_id", None)
    if not is_topic_allowed(m.chat.id, thread_id):
        await m.answer("⚠️ Эта команда доступна только в конкретных темах (топиках). Обратитесь к администратору.")
        return
    await state.set_state(Info.waiting_object)
    await state.update_data(
        owner_id=m.from_user.id,
        work_chat_id=m.chat.id,
        work_thread_id=thread_id,
    )
    await m.answer("📝 Введите один или несколько номеров объектов (через запятую):")

@router.message(Command("photo"))
async def cmd_photo(m: Message, state: FSMContext):
    thread_id = getattr(m, "message_thread_id", None)
    if not is_topic_allowed(m.chat.id, thread_id):
        await m.answer("⚠️ Эта команда доступна только в конкретных темах (топиках). Обратитесь к администратору.")
        return
    await state.set_state(Upload.waiting_object)
    await state.update_data(
        owner_id=m.from_user.id,
        work_chat_id=m.chat.id,
        work_thread_id=thread_id,
        step_msg=None,
        save_shown_for_step=-1
    )
    await m.answer("🔢 Введите номер объекта:")

@router.message(Command("result"))
async def cmd_result(m: Message):
    rows = list_completed()
    if not rows:
        await m.answer("📋 Пока нет завершённых объектов (через /photo).", reply_markup=main_kb())
        return

    is_private_chat = m.chat.type == "private"
    current_group_chat_id = None
    
    if not is_private_chat:
        # Проверяем, привязан ли Excel к этому групповому чату
        filename = get_excel_filename_for_chat(m.chat.id)
        if filename:
            current_group_chat_id = str(m.chat.id) # Cохраняем как строку для согласованности
        else:
            # Если команда в группе, но БЕЗ Excel, ведем себя как в ЛС (показываем все)
            is_private_chat = True 

    # --- Построение карты {object_id -> group_id} ---
    # Это делается один раз за команду, читая все Excel-файлы.
    logger.info("Построение карты object_id -> group_id для /result...")
    obj_to_group_map = {}
    local_excel_map = stringify_keys(EXCEL_MAP) # Гарантируем строковые ключи
    
    for chat_id_str, filename in local_excel_map.items():
        try:
            wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
            sh = wb.active
            for row in sh.iter_rows(min_row=2, values_only=True):
                if row and row[0]: # Проверяем, что есть ID
                    obj_id = str(row[0]).strip()
                    if obj_id:
                        obj_to_group_map[obj_id] = chat_id_str
        except Exception as e:
            logger.warning("Ошибка при чтении %s для /result: %s", filename, e)
    logger.info("Карта для /result построена.")
    # --- Карта построена ---

    lines = ["✅ Обработанные объекты:"]
    
    # --- Логика для приватного чата (группировка) ---
    if is_private_chat:
        objects_by_group = {} # { "group_id_str": [ "строка_объекта", ... ] }
        
        for oid, author, ts in rows:
            try:
                ts_h = datetime.fromisoformat(ts).strftime("%d.%m.%Y %H:%M")
            except:
                ts_h = ts
            
            # Находим группу для объекта
            group_id = obj_to_group_map.get(oid) or "Неизвестная группа" 
            
            if group_id not in objects_by_group:
                objects_by_group[group_id] = []
            objects_by_group[group_id].append(f"• #{oid} — {author} ({ts_h})")

        if not objects_by_group:
             await m.answer("📋 Пока нет завершённых объектов (через /photo).", reply_markup=main_kb())
             return

        # Сортируем группы (по ID) для предсказуемого порядка
        sorted_group_ids = sorted(objects_by_group.keys())

        for group_id in sorted_group_ids:
            items = objects_by_group[group_id]
            group_name = group_id
            if group_id != "Неизвестная группа":
                 # Пытаемся получить имя файла для наглядности
                 fname = local_excel_map.get(group_id) 
                 if fname:
                     group_name = f"{group_id} ({fname})"
             
            lines.append(f"\n📁 Группа: {group_name}")
            lines.extend(items)

    # --- Логика для группового чата (фильтрация) ---
    else:
        filtered_items = []
        for oid, author, ts in rows:
            group_id = obj_to_group_map.get(oid)
            
            # Показываем только если объект принадлежит ТЕКУЩЕЙ группе
            if group_id == current_group_chat_id:
                try:
                    ts_h = datetime.fromisoformat(ts).strftime("%d.%m.%Y %H:%M")
                except:
                    ts_h = ts
                filtered_items.append(f"• #{oid} — {author} ({ts_h})")
        
        if not filtered_items:
            await m.answer("📋 В этой группе пока нет завершённых объектов (через /photo).", reply_markup=main_kb())
            return
        
        lines.extend(filtered_items)
    
    # --- Отправка ответа (с разбивкой, если нужно) ---
    message_text = "\n".join(lines)
    if len(message_text) > 4096:
        logger.warning("Сообщение /result слишком длинное (%d), будет разбито.", len(message_text))
        # Отправляем первую часть с клавиатурой
        await m.answer(message_text[:4096], reply_markup=main_kb())
        # Отправляем остальные части без
        for i in range(4096, len(message_text), 4096):
            await m.answer(message_text[i:i+4096])
    else:
        await m.answer(message_text, reply_markup=main_kb())

# ========== CHECK OBJECT & CONFIRM ==========
@router.message(Upload.waiting_object)
async def check_upload_object(m: Message, state: FSMContext):
    obj = m.text.strip()
    ok, name = check_object_excel(m.chat.id, obj)
    if ok:
        # ensure topic is allowed still (race safety)
        if not is_topic_allowed(m.chat.id, getattr(m, "message_thread_id", None)):
            await m.answer("⚠️ Эта тема не привязана к маршруту. Операция отменена.")
            await state.clear()
            return
        await state.update_data(
            object=obj,
            object_name=name,
            step=0,
            steps=[{"name": s, "files": []} for s in UPLOAD_STEPS],
            owner_id=m.from_user.id,
            work_chat_id=m.chat.id,
            work_thread_id=getattr(m, "message_thread_id", None),
            step_msg=None,
            save_shown_for_step=-1
        )
        await state.set_state(Upload.confirming)
        await m.answer(
            f"Подтвердите объект:\n\n🆔 {obj}\n🏷️ {name}",
            reply_markup=confirm_kb("photo", m.from_user.id)
        )
    elif ok is False:
        await m.answer(f"❌ Объект {obj} не найден.")
        await state.clear()
    else:
        await m.answer(f"❌ Ошибка чтения: {name or 'неизвестно'}")
        await state.clear()

@router.message(AddPhoto.waiting_object)
async def check_add_object(m: Message, state: FSMContext):
    obj = m.text.strip()
    ok, name = check_object_excel(m.chat.id, obj)
    if ok:
        # ensure topic allowed
        if not is_topic_allowed(m.chat.id, getattr(m, "message_thread_id", None)):
            await m.answer("⚠️ Эта тема не привязана к маршруту. Операция отменена.")
            await state.clear()
            return
        await state.update_data(
            object=obj,
            object_name=name,
            files=[],
            owner_id=m.from_user.id,
            work_chat_id=m.chat.id,
            work_thread_id=getattr(m, "message_thread_id", None),
            step_msg=None
        )
        await state.set_state(AddPhoto.confirming)
        await m.answer(
            f"Подтвердите объект (добавление файлов):\n\n🆔 {obj}\n🏷️ {name}",
            reply_markup=confirm_kb("add", m.from_user.id)
        )
    elif ok is False:
        await m.answer(f"❌ Объект {obj} не найден.")
        await state.clear()
    else:
        await m.answer(f"❌ Ошибка чтения: {name or 'неизвестно'}")
        await state.clear()

# ===== Confirmations =====
@router.callback_query(F.data.startswith("photo_confirm_yes_"))
async def photo_confirm_yes(c: CallbackQuery, state: FSMContext):
    user_id = int(c.data.split("_")[-1])
    if c.from_user.id != user_id:
        await safe_cq_answer(c, "Эта кнопка не для вас 😅", show_alert=True)
        return

    data = await state.get_data()
    # verify topic still allowed
    if not is_topic_allowed(data.get("work_chat_id"), data.get("work_thread_id")):
        await safe_cq_answer(c, "⚠️ Тема больше не привязана к маршруту. Операция отменена.", show_alert=True)
        await state.clear()
        return

    await state.set_state(Upload.uploading)
    data = await state.get_data()
    step0 = data["steps"][0]["name"]
    owner_id = data.get("owner_id")
    try:
        await safe_edit_message(c.message, step0, reply_markup=cancel_only_kb(owner_id))
        await state.update_data(step_msg=(c.message.chat.id, c.message.message_id), save_shown_for_step=-1)
    except Exception:
        sent = await safe_call(bot.send_message(data["work_chat_id"], step0, reply_markup=cancel_only_kb(owner_id), message_thread_id=data.get("work_thread_id")))
        if sent:
            await state.update_data(step_msg=(sent.chat.id, sent.message_id), save_shown_for_step=-1)
    await safe_cq_answer(c, "Подтверждено ✅")

@router.callback_query(F.data.startswith("add_confirm_yes_"))
async def add_confirm_yes(c: CallbackQuery, state: FSMContext):
    user_id = int(c.data.split("_")[-1])
    if c.from_user.id != user_id:
        await safe_cq_answer(c, "Эта кнопка не для вас 😅", show_alert=True)
        return

    data = await state.get_data()
    # verify topic allowed
    if not is_topic_allowed(data.get("work_chat_id"), data.get("work_thread_id")):
        await safe_cq_answer(c, "⚠️ Тема не привязана к маршруту. Операция отменена.", show_alert=True)
        await state.clear()
        return

    obj = data.get("object")
    owner_id = data.get("owner_id")
    await state.set_state(AddPhoto.uploading)
    try:
        await safe_edit_message(c.message, f"📸 Отправьте дополнительные файлы для объекта №{obj}.", reply_markup=cancel_only_kb(owner_id))
        await state.update_data(step_msg=(c.message.chat.id, c.message.message_id))
    except Exception:
        sent = await safe_call(bot.send_message(data["work_chat_id"], f"📸 Отправьте дополнительные файлы для объекта №{obj}.", reply_markup=cancel_only_kb(owner_id), message_thread_id=data.get("work_thread_id")))
        if sent:
            await state.update_data(step_msg=(sent.chat.id, sent.message_id))
    await safe_cq_answer(c, "Подтверждено ✅")

# ===== Cancel (universal) =====
@router.callback_query(F.data.startswith("cancel_"))
async def cancel_anywhere(c: CallbackQuery, state: FSMContext):
    parts = c.data.split("_", 1)
    if len(parts) < 2:
        await safe_cq_answer(c, "Эта кнопка не для вас 😅", show_alert=True)
        return
    try:
        user_id = int(parts[1])
    except:
        await safe_cq_answer(c, "Эта кнопка не для вас 😅", show_alert=True)
        return

    if c.from_user.id != user_id:
        await safe_cq_answer(c, "Эта кнопка не для вас 😅", show_alert=True)
        return

    await state.clear()
    try:
        await safe_edit_message(c.message, "❌ Действие отменено.", reply_markup=None)
    except:
        pass
    await safe_cq_answer(c, "Отменено ✅")

# ========== MEDIA HANDLING ==========
async def _finalize_media_group_for_photo(m: Message, state: FSMContext, group_id: str):
    await asyncio.sleep(3.2)
    data = await state.get_data()
    step_i = data.get("step", 0)
    steps = data.get("steps", [])
    if step_i >= len(steps):
        return
    cur = steps[step_i]

    media_groups = data.get("media_groups", {})
    finalizing = set(data.get("finalizing_groups", []))

    if group_id not in media_groups:
        return
    if group_id in finalizing:
        return

    finalizing.add(group_id)
    await state.update_data(finalizing_groups=list(finalizing))

    files = media_groups[group_id]
    cur["files"].extend(files)

    await state.update_data(steps=steps, media_groups=media_groups)

    # показать пользователю действие — заменяя предыдущее сообщение шага
    await _show_action_for_current_step(state, m.chat.id, step_i, m)

    finalizing.discard(group_id)
    await state.update_data(finalizing_groups=list(finalizing))

async def _finalize_media_group_for_addphoto(m: Message, state: FSMContext, group_id: str):
    await asyncio.sleep(3.2)
    data = await state.get_data()
    files = data.get("files", [])

    media_groups = data.get("media_groups", {})
    finalizing = set(data.get("finalizing_groups", []))

    if group_id not in media_groups:
        return
    if group_id in finalizing:
        return

    finalizing.add(group_id)
    await state.update_data(finalizing_groups=list(finalizing))

    files.extend(media_groups[group_id])
    await state.update_data(files=files, media_groups=media_groups)

    # показать action_kb для addphoto
    await _show_action_for_addphoto(state, m.chat.id, m)

    finalizing.discard(group_id)
    await state.update_data(finalizing_groups=list(finalizing))

async def _handle_media_group(m: Message, state: FSMContext, group_id: str, handler):
    data = await state.get_data()
    media_groups = data.get("media_groups", {})
    if group_id not in media_groups:
        media_groups[group_id] = []
    media_groups[group_id].append({
        "type": "photo" if m.photo else "video" if m.video else "document",
        "file_id": m.photo[-1].file_id if m.photo else m.video.file_id if m.video else m.document.file_id,
    })
    await state.update_data(media_groups=media_groups)
    asyncio.create_task(handler(m, state, group_id))

@router.message(Upload.uploading, F.media_group_id)
async def album_uploading(m: Message, state: FSMContext):
    await _handle_media_group(m, state, m.media_group_id, _finalize_media_group_for_photo)

@router.message(AddPhoto.uploading, F.media_group_id)
async def album_addphoto(m: Message, state: FSMContext):
    await _handle_media_group(m, state, m.media_group_id, _finalize_media_group_for_addphoto)

@router.message(Upload.uploading, (F.photo | F.video | F.document) & ~F.media_group_id)
async def single_file_uploading(m: Message, state: FSMContext):
    data = await state.get_data()
    step_i = data.get("step", 0)
    steps = data.get("steps", [])
    if step_i >= len(steps):
        await m.answer("⚠️ Ошибка состояния: этап загрузки не найден. Попробуйте /photo заново.")
        await state.clear()
        return
    cur = steps[step_i]
    file_type = "photo" if m.photo else "video" if m.video else "document"
    file_id = m.photo[-1].file_id if m.photo else m.video.file_id if m.video else m.document.file_id
    cur["files"].append({"type": file_type, "file_id": file_id})
    await state.update_data(steps=steps)
    # показать action клавиатуру "Выберите действие:"
    await _show_action_for_current_step(state, m.chat.id, step_i, m)

@router.message(AddPhoto.uploading, (F.photo | F.video | F.document) & ~F.media_group_id)
async def single_file_addphoto(m: Message, state: FSMContext):
    data = await state.get_data()
    files = data.get("files", [])
    file_type = "photo" if m.photo else "video" if m.video else "document"
    file_id = m.photo[-1].file_id if m.photo else m.video.file_id if m.video else m.document.file_id
    files.append({"type": file_type, "file_id": file_id})
    await state.update_data(files=files)
    # показать action клавиатуру для addphoto
    await _show_action_for_addphoto(state, m.chat.id, m)

# show "Выберите действие:" for current step in /photo flow
async def _show_action_for_current_step(state: FSMContext, chat_id: int, step_i: int, context_msg: Message):
    data = await state.get_data()
    save_shown_for_step = data.get("save_shown_for_step", -1)
    if save_shown_for_step == step_i:
        return
    # delete previous step message if exists
    step_msg = data.get("step_msg")
    if step_msg:
        try:
            stored_chat, stored_mid = step_msg
            await safe_call(bot.delete_message(stored_chat, stored_mid))
        except Exception:
            logger.debug("Could not delete previous step message", exc_info=True)
    # send action message "Выберите действие:" 
    owner_id = data.get("owner_id")
    work_thread_id = data.get("work_thread_id")
    try:
        sent = await safe_call(bot.send_message(chat_id, "Выберите действие:", reply_markup=action_kb(owner_id), message_thread_id=work_thread_id))
        if sent:
            await state.update_data(step_msg=(sent.chat.id, sent.message_id), save_shown_for_step=step_i)
    except Exception:
        logger.exception("Failed to send action message for current step")

# show "Выберите действие:" for addphoto flow
async def _show_action_for_addphoto(state: FSMContext, chat_id: int, context_msg: Message):
    data = await state.get_data()
    step_msg = data.get("step_msg")
    if step_msg:
        try:
            stored_chat, stored_mid = step_msg
            await safe_call(bot.delete_message(stored_chat, stored_mid))
        except Exception:
            logger.debug("Could not delete previous addphoto message", exc_info=True)
    owner_id = data.get("owner_id")
    work_thread_id = data.get("work_thread_id")
    try:
        sent = await safe_call(bot.send_message(chat_id, "Выберите действие:", reply_markup=action_kb(owner_id), message_thread_id=work_thread_id))
        if sent:
            await state.update_data(step_msg=(sent.chat.id, sent.message_id))
    except Exception:
        logger.exception("Failed to send action message for addphoto")

# ========== SAVE CALLBACK (handles both /photo and addphoto save) ==========
@router.callback_query(F.data.startswith("save_"))
async def save_callback(c: CallbackQuery, state: FSMContext):
    # save_{user_id} — user_id may be empty string
    try:
        uid_part = c.data.split("_", 1)[1]
        user_id = int(uid_part) if uid_part != "" else None
    except:
        user_id = None

    data = await state.get_data()
    # verify user if user_id present
    if user_id and c.from_user.id != user_id:
        await safe_cq_answer(c, "Эта кнопка не для вас 😅", show_alert=True)
        return

    # Photo flow
    if "steps" in data:
        step_i = data.get("step", 0)
        steps = data.get("steps", [])
        if not isinstance(steps, list) or step_i >= len(steps):
            await safe_cq_answer(c)
            await state.clear()
            return
        cur = steps[step_i]
        if not cur.get("files"):
            await safe_cq_answer(c)
            return
        obj = data.get("object")
        # save step to DB
        save_files(obj, cur["name"], cur["files"], c.from_user.full_name)
        # advance
        step_i += 1
        await state.update_data(step=step_i, steps=steps, save_shown_for_step=-1)
        # if finished all steps -> background archive + move to video state
        if step_i >= len(steps):
            work_chat_id = data.get("work_chat_id")
            work_thread_id = data.get("work_thread_id")
            route = mapping_lookup(work_chat_id, work_thread_id)
            owner_id = data.get("owner_id")
            obj_name = data.get("object_name")
            steps_copy = steps.copy()
            if route:
                archive_chat_id = route["chat_id"]
                archive_thread_id = route["thread_id"]
                asyncio.create_task(_archive_and_notify(owner_id, obj, obj_name, steps_copy, archive_chat_id, archive_thread_id, c.from_user.full_name))
            else:
                logger.warning("No route for work_chat_id=%s thread=%s", work_chat_id, work_thread_id)
                try:
                    await safe_call(bot.send_message(owner_id, "⚠️ В этой теме не настроен маршрут в Архив. Обратитесь к разработчику."))
                except Exception:
                    logger.exception("Could not notify owner about missing route")
            await state.set_state(Upload.uploading_video)
            # edit message to show video step prompt with cancel button and store it
            try:
                await safe_edit_message(c.message, VIDEO_STEP, reply_markup=cancel_only_kb(owner_id))
                await state.update_data(step_msg=(c.message.chat.id, c.message.message_id))
            except Exception:
                sent = await safe_call(bot.send_message(data["work_chat_id"], VIDEO_STEP, reply_markup=cancel_only_kb(owner_id), message_thread_id=data.get("work_thread_id")))
                if sent:
                    await state.update_data(step_msg=(sent.chat.id, sent.message_id))
            return
        # else show next step initial message (only Cancel)
        next_step = steps[step_i]["name"]
        owner_id = data.get("owner_id")
        try:
            await safe_edit_message(c.message, next_step, reply_markup=cancel_only_kb(owner_id))
            await state.update_data(step_msg=(c.message.chat.id, c.message.message_id))
        except Exception:
            sent = await safe_call(bot.send_message(data["work_chat_id"], next_step, reply_markup=cancel_only_kb(owner_id), message_thread_id=data.get("work_thread_id")))
            if sent:
                await state.update_data(step_msg=(sent.chat.id, sent.message_id))
        await safe_cq_answer(c, "✅ Сохранено")
        return

    # Addphoto flow
    elif "files" in data:
        files = data.get("files", [])
        if not files:
            await safe_cq_answer(c)
            return
        obj = data.get("object")
        obj_name = data.get("object_name")
        owner_id = data.get("owner_id")
        work_chat_id = data.get("work_chat_id")
        work_thread_id = data.get("work_thread_id")
        # save to DB
        save_files(obj, "Дополнительные фото", files, c.from_user.full_name)
        # Try to edit action message into confirmation text
        try:
            await safe_edit_message(c.message, "✅ Файлы отправлены. Спасибо!", reply_markup=None)
        except Exception:
            logger.debug("Could not edit action message to confirmation; will try to send a confirmation message")
            try:
                if c.message:
                    await safe_call(bot.delete_message(c.message.chat.id, c.message.message_id))
            except Exception:
                pass
            try:
                await safe_call(bot.send_message(owner_id, "✅ Файлы отправлены. Спасибо!"))
            except Exception:
                pass
        # send to archive in background using stored route
        route = mapping_lookup(work_chat_id, work_thread_id)
        if route:
            archive_chat_id = route["chat_id"]
            archive_thread_id = route["thread_id"]
            asyncio.create_task(_send_header_and_files_to_archive(obj, obj_name, files, archive_chat_id, archive_thread_id, c.from_user.full_name))
        else:
            logger.warning("No route for addphoto: work_chat_id=%s thread=%s", work_chat_id, work_thread_id)
            try:
                await safe_call(bot.send_message(owner_id, "⚠️ В этой теме не настроен маршрут в Архив. Обратитесь к разработчику."))
            except Exception:
                logger.exception("Could not notify owner about missing route")
        await state.clear()
        await safe_cq_answer(c, "✅ Файлы отправлены. Идёт отправка в Архив.")
        return

    else:
        await safe_cq_answer(c)
        return

# ========== BACKGROUND ARCHIVE & NOTIFY ==========
async def _archive_and_notify(owner_id: int, obj: str, obj_name: str, steps: list, chat_id: int, thread_id: int, author: str):
    try:
        # Send header first
        header_text = f"Объект #{obj}\n🏠 {obj_name}\n🙋🏻‍♂️ {author}"
        try:
            await safe_call(bot.send_message(chat_id, header_text, message_thread_id=thread_id))
        except Exception:
            logger.exception("Failed to send header to archive")
        # then send files grouped (by 10)
        media_buffer = []
        for step in steps:
            for f in step.get("files", []):
                if f["type"] == "photo":
                    media_buffer.append(InputMediaPhoto(media=f["file_id"]))
                elif f["type"] == "video":
                    media_buffer.append(InputMediaVideo(media=f["file_id"]))
                else:
                    media_buffer.append(InputMediaDocument(media=f["file_id"]))
                if len(media_buffer) >= 10:
                    await safe_call(bot.send_media_group(chat_id, media_buffer, message_thread_id=thread_id))
                    media_buffer = []
        if media_buffer:
            await safe_call(bot.send_media_group(chat_id, media_buffer, message_thread_id=thread_id))
        # notify user — removed per request (do not send notification to owner)
        # (original code sent a message to owner_id here; it is intentionally removed)
    except Exception:
        logger.exception("Error during background archive")
        try:
            await safe_call(bot.send_message(owner_id, "⚠️ Ошибка при отправке в архив."))
        except Exception:
            pass

# helper: send header + provided files list to archive (used for addphoto single-step)
async def _send_header_and_files_to_archive(obj: str, obj_name: str, files: list, chat_id: int, thread_id: int, author: str):
    try:
        header_text = f"Объект #{obj}\n🏠 {obj_name}\n🙋🏻‍♂️ {author}"
        try:
            await safe_call(bot.send_message(chat_id, header_text, message_thread_id=thread_id))
        except Exception:
            logger.exception("Failed to send header to archive (addphoto)")
        media_buffer = []
        for f in files:
            if f["type"] == "photo":
                media_buffer.append(InputMediaPhoto(media=f["file_id"]))
            elif f["type"] == "video":
                media_buffer.append(InputMediaVideo(media=f["file_id"]))
            else:
                media_buffer.append(InputMediaDocument(media=f["file_id"]))
            if len(media_buffer) >= 10:
                await safe_call(bot.send_media_group(chat_id, media_buffer, message_thread_id=thread_id))
                media_buffer = []
        if media_buffer:
            await safe_call(bot.send_media_group(chat_id, media_buffer, message_thread_id=thread_id))
    except Exception:
        logger.exception("Error sending addphoto files to archive")

# ========== VIDEO UPLOAD ==========
@router.message(Upload.uploading_video, (F.video | F.document))
async def video_uploading(m: Message, state: FSMContext):
    data = await state.get_data()
    obj = data.get("object")
    obj_name = data.get("object_name")
    work_chat_id = data.get("work_chat_id")
    work_thread_id = data.get("work_thread_id")
    owner_id = data.get("owner_id")
    route = mapping_lookup(work_chat_id, work_thread_id)
    if not route:
        await m.answer("⚠️ В этой теме не настроен маршрут в Архив. Обратитесь к разработчику.")
        await state.clear()
        return

    archive_chat_id = route["chat_id"]
    archive_thread_id = route["thread_id"]

    file_id = None
    if m.video:
        file_id = m.video.file_id
    elif m.document and getattr(m.document, "mime_type", "").startswith("video"):
        file_id = m.document.file_id
    else:
        await m.answer("Пожалуйста, пришлите видео или видеофайл.")
        return

    try:
        # отправка видео в архив сразу
        await safe_call(bot.send_video(archive_chat_id, file_id, message_thread_id=archive_thread_id))
    except Exception as e:
        logger.exception("Error sending video to archive: %s", e)
        await m.answer(f"⚠️ Ошибка при отправке видео в архив: {e}")
        return

    save_files(obj, VIDEO_STEP, [{"type": "video", "file_id": file_id}], m.from_user.full_name)
    mark_completed(obj, m.from_user.full_name)

    # удаляем сообщение шага для видео (если есть)
    step_msg = data.get("step_msg")
    if step_msg:
        try:
            stored_chat, stored_mid = step_msg
            await safe_call(bot.delete_message(stored_chat, stored_mid))
        except Exception:
            logger.debug("Could not delete video step message", exc_info=True)

    await m.answer("✅ Видео отправлено и объект завершён. Спасибо!")
    await state.clear()

# ========== INFO ==========
@router.message(Info.waiting_object)
async def info_objects(m: Message, state: FSMContext):
    objs = [o.strip() for o in m.text.split(",")]
    lines = []
    for obj in objs:
        info = get_object_info(m.chat.id, obj)
        if isinstance(info, dict) and "error" in info:
            lines.append(f"❌ #{obj}: {info['error']}")
        elif info is None:
            lines.append(f"❌ #{obj}: не найден")
        else:
            lines.append(
                f"✅ #{info['id']}\n"
                f"🏷️ {info['consumer']}\n"
                f"🏢 {info['object']}\n"
                f"📍 {info['address']}"
            )
    await m.answer("\n\n".join(lines) if lines else "—", reply_markup=main_kb())
    await state.clear()

# ========== WEBHOOK HANDLER & SHUTDOWN ==========
async def webhook_handler(request: web.Request):
    try:
        data = await request.json()
    except Exception:
        return web.Response(status=400, text="invalid json")
    asyncio.create_task(dp.feed_webhook_update(bot, data))
    return web.Response(text="ok")

async def on_shutdown(app):
    try:
        await bot.session.close()
    except Exception:
        logger.exception("Error closing bot.session on shutdown")

# ========== MAIN ==========
async def main():
    init_db()
    load_settings()
    dp.include_router(router)

    commands = [
        BotCommand(command="start", description="Запуск бота"),
        BotCommand(command="addphoto", description="Добавить фото к объекту"),
        BotCommand(command="info", description="Информация по объектам"),
        BotCommand(command="photo", description="Новая загрузка"),
        BotCommand(command="result", description="Список обработанных объектов")
    ]
    try:
        await bot.set_my_commands(commands)
    except Exception:
        logger.exception("Failed to set bot commands")

    asyncio.create_task(keepalive())

    try:
        await bot.set_webhook(f"{WEBHOOK_URL}/webhook")
    except Exception:
        logger.exception("Failed to set webhook")

    app = web.Application()
    app.router.add_post("/webhook", webhook_handler)
    app.on_cleanup.append(on_shutdown)
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", PORT)
    await site.start()
    logger.info(f"🚀 Бот запущен на порту {PORT}")

    while True:
        await asyncio.sleep(3600)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except (KeyboardInterrupt, SystemExit):
        logger.info("Shutting down")
